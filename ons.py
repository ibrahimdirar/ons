import requests
import tkinter as tk
import os
import sys
from win32com.client import Dispatch
import openpyxl as xl
import datetime
import numpy as np
import pandas as pd

months = ['January', 'February', 'March', 'April', 'May', 'June',
          'July', 'August', 'September', 'October', 'November', 'December']


def print_table(title, dataset, row):
    ws.cell(row=row, column=1, value=title)
    col = 1
    for month in months:
        col += 1
        ws.cell(row=row, column=col, value=month[:3])
    col = 1
    row += 1
    for data in dataset:
        if(data['month'] == 'January'):
            ws.cell(row=row, column=1, value=int(data['year']))
        col += 1
        ws.cell(row=row, column=col, value=float(
            data['value'])).number_format = '0.0'
        if(data['month'] == 'December'):
            col = 1
            row += 1
    print_table.last_row = row
    # col = 1
    # for data in dataset:
    #     if(data['month'] == 'January'):
    #         col += 1
    #         row = orow + 1
    #         ws.cell(row=row - 1, column=col, value=data['year'])
    #     else:
    #         row += 1
    #     ws.cell(row=row, column=col, value=data['value'])


root = tk.Tk()
root.withdraw()

wb = xl.Workbook()
ws = wb.active

r_rpi = requests.get(
    "https://api.ons.gov.uk/dataset/mm23/timeseries/CHAW/data")
r_cpi = requests.get(
    "https://api.ons.gov.uk/dataset/mm23/timeseries/D7BT/data")
r_rpix = requests.get(
    "https://api.ons.gov.uk/dataset/mm23/timeseries/CHMK/data")

row = 1

current_year = now = datetime.datetime.now().year
start_year = 2016


print_table('RPI', r_rpi.json()['months'], row)

row_offset = 3


print_table('CPI', r_cpi.json()['months'],
            print_table.last_row + row_offset + 3)

print_table('RPIX', r_rpix.json()['months'], print_table.last_row + 3)

file_name = 'inflationandpriceindices.xlsx'
wb.save(filename=file_name)

excel = Dispatch('Excel.Application')
excel.Visible = True

# Opens the excel file, bundled version will have sys.frozen set to True
if getattr(sys, 'frozen', False):
    # running in a bundle
    # os.system('start excel.exe "%s\\output.csv"' % (sys._MEIPASS, ))
    wb = excel.Workbooks.Open(
        "%s\\inflationandpriceindices.xlsx" % (os.path.abspath("."), ))
else:
    # running live
    wb = excel.Workbooks.Open(
        "%s\\inflationandpriceindices.xlsx" % (sys.path[0], ))

excel.Worksheets(1).Activate()
excel.ActiveSheet.Columns.AutoFit()

wb.Save()


# if (r_rpi.status_code != 404):
#     for data in r_rpi.json()['months']:
#         print('RPI: ', data['date'], ' - ', data['value'])
#     for data in r_cpi.json()['months']:
#         print('CPI: ', data['date'], ' - ', data['value'])
#     for data in r_rpix.json()['months']:
#         print('RPIX: ', data['date'], ' - ', data['value'])
